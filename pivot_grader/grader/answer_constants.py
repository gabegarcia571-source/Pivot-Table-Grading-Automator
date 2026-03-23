"""Answer-key derived constants used by question-specific graders."""

from __future__ import annotations

from pathlib import Path


def load_holiday_customers_from_answer_key(answer_key_path: str) -> frozenset[int]:
    """Read Q8 in the answer key and return IDs with purchases only in Nov/Dec."""
    try:
        import openpyxl
    except Exception:
        return frozenset()

    workbook_path = Path(answer_key_path)
    if not workbook_path.exists():
        return frozenset()

    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception:
        return frozenset()

    if "Q8" not in wb.sheetnames:
        wb.close()
        return frozenset()

    ws = wb["Q8"]
    ids: set[int] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        customer_id = row[0]
        if customer_id is None:
            continue

        months = row[1:13]  # Jan..Dec
        jan_oct = months[:10]
        nov_dec = months[10:12]

        has_jan_oct = any(v is not None and str(v).strip() != "" for v in jan_oct)
        has_nov_dec = any(v is not None and str(v).strip() != "" for v in nov_dec)

        if (not has_jan_oct) and has_nov_dec:
            try:
                ids.add(int(customer_id))
            except (TypeError, ValueError):
                continue

    wb.close()
    return frozenset(ids)


def _default_answer_key_path() -> Path:
    return Path(__file__).resolve().parents[1] / "answer_key" / "GryzzlSales2024 - Answer Key.xlsx"

# Q8 — customers who order ONLY in November and/or December.
HOLIDAY_ONLY_CUSTOMERS: frozenset[int] = load_holiday_customers_from_answer_key(
    str(_default_answer_key_path())
)

# Safety flag for graders: True only when the extracted set is complete.
HOLIDAY_ONLY_CUSTOMERS_COMPLETE = len(HOLIDAY_ONLY_CUSTOMERS) == 2487
