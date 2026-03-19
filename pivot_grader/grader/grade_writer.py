from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

from grader.scoring import format_short_comments


def write_grades(
    student_id: str,
    scores_dict: dict[str, float | None],
    comments_dict: dict[str, str],
    template_path: str | Path,
    output_dir: str | Path,
    max_comment_words: int = 15,
) -> Path:
    """Write student grading results into a template copy."""
    template = Path(template_path)
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    output_path = out_dir / f"{student_id}_grade.xlsx"
    shutil.copy2(template, output_path)

    wb = load_workbook(output_path)
    ws = wb["Gradesheet"]

    ws["C3"] = student_id

    for i in range(1, 11):
        qid = f"Q{i}"
        row = 6 + i  # Q1->7 ... Q10->16
        score = scores_dict.get(qid)
        ws[f"C{row}"] = "" if score is None else float(score)
        raw_comment = comments_dict.get(qid, "")
        ws[f"D{row}"] = format_short_comments([raw_comment], max_words=max_comment_words)

    wb.save(output_path)
    return output_path
