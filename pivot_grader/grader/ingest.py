from __future__ import annotations

import os
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


def _trace_enabled() -> bool:
    return os.getenv("PIVOT_TRACE_MODE", "0").strip().lower() in {"1", "true", "yes", "on"}


def _trace(step: str, message: str = "") -> None:
    if not _trace_enabled():
        return
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds")
    suffix = f" | {message}" if message else ""
    print(f"[TRACE {ts}] {step}{suffix}")


@dataclass
class SubmissionResult:
    student_id: str
    sheets: dict[str, pd.DataFrame]
    error: str | None = None
    workbook_path: Path | None = None


def _normalize_frame(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = df.dropna(how="all").dropna(axis=1, how="all")
    cleaned.columns = [str(col).strip() for col in cleaned.columns]
    return cleaned


_STRICT_URIS: tuple[tuple[bytes, bytes], ...] = (
    # Strict OOXML spreadsheet namespace -> Transitional namespace
    (
        b"http://purl.oclc.org/ooxml/spreadsheetml/main",
        b"http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    ),
    # Strict OOXML officeDocument relationships namespace -> Transitional
    (
        b"http://purl.oclc.org/ooxml/officeDocument/relationships",
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    ),
)


def _is_strict_ooxml(path: Path) -> bool:
    """Return True when xl/workbook.xml uses Strict OOXML namespaces."""
    try:
        with zipfile.ZipFile(path) as zf:
            workbook_xml = zf.read("xl/workbook.xml")
    except Exception:  # noqa: BLE001
        return False
    return any(old in workbook_xml for old, _ in _STRICT_URIS)


def _convert_strict_ooxml_to_transitional(path: Path) -> Path:
    """Create a transitional-namespace copy of a strict OOXML .xlsx file."""
    tmp = tempfile.NamedTemporaryFile(prefix="pivot_fix_", suffix=".xlsx", delete=False)
    tmp_path = Path(tmp.name)
    tmp.close()

    with zipfile.ZipFile(path, "r") as src, zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            payload = src.read(info.filename)
            if info.filename.endswith(".xml"):
                for old, new in _STRICT_URIS:
                    payload = payload.replace(old, new)
            dst.writestr(info, payload)

    return tmp_path


def find_pivot_origin(ws) -> tuple[int, int]:
    """
    Scan an openpyxl worksheet for the first non-empty string cell in a row
    that has at least 2 non-empty values — more likely a table header than a
    stray label.  Falls back to the very first non-empty string cell found.
    Returns (row, col) as 1-indexed openpyxl coordinates.
    """
    first_string_pos: tuple[int, int] | None = None

    for row_cells in ws.iter_rows():
        non_empty_in_row = sum(
            1 for cell in row_cells
            if cell.value is not None and str(cell.value).strip()
        )
        for cell in row_cells:
            if cell.value and isinstance(cell.value, str) and cell.value.strip():
                if first_string_pos is None:
                    first_string_pos = (cell.row, cell.column)
                # Prefer a row with ≥2 non-empty cells — much more likely to be a header
                if non_empty_in_row >= 2:
                    return cell.row, cell.column

    return first_string_pos or (1, 1)


def load_answer_key(path: str | Path) -> dict[str, pd.DataFrame]:
    """Load answer key workbook with header on row 3 (skip first 2 blank rows)."""
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Answer key not found: {workbook_path}")

    sheets = pd.read_excel(workbook_path, sheet_name=None, header=2, engine="openpyxl")
    return {name: _normalize_frame(df) for name, df in sheets.items()}


def load_student_submission(folder_path: str | Path) -> SubmissionResult:
    """Load all sheets from the first xlsx file found in the student folder.

    For each sheet, openpyxl is used in read-only/stream mode to locate the
    pivot table origin (the first header-like row).  pandas then reads from
    that row onward, and the DataFrame is trimmed to the pivot's starting
    column so that students who offset their tables still compare correctly.
    """
    path = Path(folder_path)
    student_id = path.stem if path.is_file() else path.name

    try:
        _trace("submission discovery", f"inspect {path}")
        if path.is_file():
            if path.suffix.lower() != ".xlsx":
                return SubmissionResult(
                    student_id=student_id,
                    sheets={},
                    error=f"Unsupported submission file type: {path.name}",
                )
            xlsx_files = [path]
        else:
            xlsx_files = sorted(
                [
                    p
                    for p in path.glob("*.xlsx")
                    if not p.name.startswith("~$") and "grade" not in p.name.lower()
                ]
            )
        if not xlsx_files:
            return SubmissionResult(student_id=student_id, sheets={}, error="No .xlsx submission found.")

        workbook = xlsx_files[0]
        workbook_to_read = workbook
        _trace("sheet loading", f"open workbook {workbook}")

        # Stream-scan once with openpyxl to find each sheet's pivot origin.
        _trace("pivot extraction", f"scan pivot origins for {workbook_to_read.name}")
        opxl_wb = openpyxl.load_workbook(workbook_to_read, read_only=True, data_only=True)
        sheet_names = list(opxl_wb.sheetnames)

        if not sheet_names and _is_strict_ooxml(workbook_to_read):
            _trace("pivot extraction", "workbook loaded with zero sheets; flagging as strict OOXML")
            opxl_wb.close()
            workbook_to_read = _convert_strict_ooxml_to_transitional(workbook)
            _trace("pivot extraction", f"strict OOXML conversion complete: {workbook_to_read}")
            opxl_wb = openpyxl.load_workbook(workbook_to_read, read_only=True, data_only=True)
            sheet_names = list(opxl_wb.sheetnames)

        origins: dict[str, tuple[int, int]] = {
            name: find_pivot_origin(opxl_wb[name]) for name in sheet_names
        }
        opxl_wb.close()

        if not origins:
            return SubmissionResult(
                student_id=student_id,
                sheets={},
                error=(
                    "Workbook loaded but has zero readable sheets. "
                    "If this file opens in Excel, save-as standard .xlsx and retry."
                ),
                workbook_path=workbook_to_read,
            )

        _trace("pivot extraction", f"found origins for {len(origins)} sheet(s)")

        sheets: dict[str, pd.DataFrame] = {}
        for sheet_name, (start_row, start_col) in origins.items():
            _trace("sheet loading", f"load sheet {sheet_name} from row={start_row}, col={start_col}")
            df = pd.read_excel(
                workbook_to_read,
                sheet_name=sheet_name,
                skiprows=start_row - 1,   # skip blank rows above the pivot
                header=0,
                engine="openpyxl",
            )
            if start_col > 1:
                df = df.iloc[:, start_col - 1:]   # trim blank columns to the left
            normalized = _normalize_frame(df.dropna(how="all"))
            # Forward-fill the first column to expand merged pivot row-group labels
            if not normalized.empty:
                normalized.iloc[:, 0] = normalized.iloc[:, 0].ffill()
            sheets[sheet_name] = normalized
        _trace("sheet loading", f"loaded {len(sheets)} sheet(s) from {workbook_to_read.name}")

        return SubmissionResult(student_id=student_id, sheets=sheets, workbook_path=workbook_to_read)
    except Exception as exc:  # noqa: BLE001
        _trace("sheet loading", f"failed loading {student_id}: {exc}")
        return SubmissionResult(student_id=student_id, sheets={}, error=f"Failed to load submission: {exc}")
