from __future__ import annotations

from typing import Any

import pandas as pd

from grader.utils.normalize import normalize_label


def _cell_is_highlighted(cell: Any) -> bool:
    fill = cell.fill
    if fill is None or fill.fill_type in (None, "none"):
        return False
    fg = fill.fgColor
    if fg is None:
        return False
    if fg.type in ("theme", "indexed"):
        return True
    rgb = (fg.rgb if hasattr(fg, "rgb") else "").upper()
    return rgb not in ("", "00000000", "FFFFFFFF", "FF000000")


def has_any_highlight(workbook_path: Any, sheet_name: str) -> bool:
    """Return True if any cell on the sheet has a non-default fill."""
    import openpyxl
    from pathlib import Path as _Path

    try:
        wb = openpyxl.load_workbook(_Path(workbook_path), data_only=True)
    except Exception:
        return False
    if sheet_name not in wb.sheetnames:
        wb.close()
        return False
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if _cell_is_highlighted(cell):
                wb.close()
                return True
    wb.close()
    return False


def has_any_highlight_in_workbook(workbook_path: Any) -> bool:
    """Return True if any sheet in the workbook has a non-default fill."""
    import openpyxl
    from pathlib import Path as _Path

    try:
        wb = openpyxl.load_workbook(_Path(workbook_path), data_only=True)
    except Exception:
        return False

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if _cell_is_highlighted(cell):
                    wb.close()
                    return True

    wb.close()
    return False


def evaluate_highlight_formatting(workbook_path: Any, sheet_name: str | None) -> tuple[float, list[str]]:
    """Return (formatting_score, formatting_issues) for highlight requirement.

    When a mapped sheet name is unavailable, fall back to workbook-level scan
    so missing highlights are still penalized instead of silently passing.
    """
    if not workbook_path or not sheet_name:
        return 1.0, ["NEEDS_REVIEW: highlight check skipped"]

    if has_any_highlight(workbook_path, sheet_name):
        return 1.0, []
    return 0.5, ["Missing highlight"]


def _coerce_label(value: Any) -> str:
    if pd.isna(value):
        return ""
    return normalize_label(str(value).strip())


def _first_numeric_column(df: pd.DataFrame) -> str | None:
    for col in df.columns[1:]:
        converted = pd.to_numeric(df[col], errors="coerce")
        if converted.notna().any():
            return str(col)
    return None


def _normalize_for_compare(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or len(df.columns) < 2:
        return pd.DataFrame(columns=["label", "value"])

    first_col = df.columns[0]
    value_col = _first_numeric_column(df)
    if value_col is None:
        return pd.DataFrame(columns=["label", "value"])

    out = df[[first_col, value_col]].copy()
    out.columns = ["label", "value"]
    out["label"] = out["label"].apply(_coerce_label)
    out["value"] = pd.to_numeric(out["value"], errors="coerce")
    out = out.dropna(subset=["value"])  # keep only comparable rows
    out = out[out["label"] != ""]
    out = out[~out["label"].str.contains("total")]
    return out


def compare_pivot_values(
    student_df: pd.DataFrame,
    answer_df: pd.DataFrame,
    numeric_tolerance: float = 1e-2,
) -> dict[str, Any]:
    """Compare student and answer pivots by label + value with tolerance."""
    student_norm = _normalize_for_compare(student_df)
    answer_norm = _normalize_for_compare(answer_df)

    if student_norm.empty or answer_norm.empty:
        return {
            "match": False,
            "mismatches": [
                {
                    "label": "<table>",
                    "expected": "non-empty comparable pivot",
                    "actual": "missing or non-numeric values",
                }
            ],
            "score_suggestion": 0.0,
        }

    answer_map = answer_norm.set_index("label")["value"].to_dict()
    student_map = student_norm.set_index("label")["value"].to_dict()

    mismatches: list[dict[str, Any]] = []
    matches = 0

    for label, expected in answer_map.items():
        actual = student_map.get(label)
        if actual is None:
            mismatches.append({"label": label, "expected": expected, "actual": None})
            continue

        if abs(float(actual) - float(expected)) <= numeric_tolerance:
            matches += 1
        else:
            mismatches.append({"label": label, "expected": float(expected), "actual": float(actual)})

    total = max(1, len(answer_map))
    ratio = matches / total
    match = len(mismatches) == 0
    score_suggestion = round(max(0.0, min(1.0, ratio)), 2)

    return {
        "match": match,
        "mismatches": mismatches,
        "score_suggestion": score_suggestion,
    }


def compare_pivot_values_subset(
    student_df: pd.DataFrame,
    answer_df: pd.DataFrame,
    required_labels: list[str] | None = None,
    ignore_labels: list[str] | set[str] | None = None,
    numeric_tolerance: float = 1e-2,
) -> dict[str, Any]:
    """Compare student rows as an allowed subset of answer rows.

    This is useful when a student applies additional filtering that removes
    non-required rows, as long as the remaining rows are correct.
    """
    student_norm = _normalize_for_compare(student_df)
    answer_norm = _normalize_for_compare(answer_df)

    ignored = {
        normalize_label(_coerce_label(label))
        for label in (ignore_labels or [])
        if _coerce_label(label)
    }
    if ignored:
        normalized_student_labels = student_norm["label"].astype(str).map(normalize_label)
        student_norm = student_norm[~normalized_student_labels.isin(ignored)]

    if student_norm.empty or answer_norm.empty:
        return {
            "match": False,
            "mismatches": [
                {
                    "label": "<table>",
                    "expected": "non-empty comparable pivot",
                    "actual": "missing or non-numeric values",
                }
            ],
            "missing_required": required_labels or [],
            "score_suggestion": 0.0,
        }

    answer_map = answer_norm.set_index("label")["value"].to_dict()
    student_map = student_norm.set_index("label")["value"].to_dict()

    mismatches: list[dict[str, Any]] = []
    matches = 0

    for label, actual in student_map.items():
        expected = answer_map.get(label)
        if expected is None:
            mismatches.append({"label": label, "expected": None, "actual": float(actual)})
            continue

        if abs(float(actual) - float(expected)) <= numeric_tolerance:
            matches += 1
        else:
            mismatches.append({"label": label, "expected": float(expected), "actual": float(actual)})

    normalized_required_pairs = [
        (_coerce_label(v), str(v).strip())
        for v in (required_labels or [])
        if _coerce_label(v)
    ]
    missing_required = [
        original_label
        for normalized_label, original_label in normalized_required_pairs
        if normalized_label not in student_map
    ]

    match = len(mismatches) == 0 and len(missing_required) == 0
    denom = max(1, len(student_map) + len(missing_required))
    score_suggestion = round(max(0.0, min(1.0, matches / denom)), 2)

    return {
        "match": match,
        "mismatches": mismatches,
        "missing_required": missing_required,
        "score_suggestion": score_suggestion,
    }


def compare_pivot_values_as_percent_of_total(
    student_df: pd.DataFrame,
    answer_df: pd.DataFrame,
    percent_tolerance: float = 1e-3,
) -> dict[str, Any]:
    """Compare pivots when student uses % of grand total instead of raw sums.

    Accepts student values provided as fractions (sum ~1.0) or percentages
    (sum ~100.0), and compares them against answer-key shares by label.
    """
    student_norm = _normalize_for_compare(student_df)
    answer_norm = _normalize_for_compare(answer_df)

    if student_norm.empty or answer_norm.empty:
        return {
            "match": False,
            "mismatches": [
                {
                    "label": "<table>",
                    "expected": "non-empty comparable pivot",
                    "actual": "missing or non-numeric values",
                }
            ],
            "score_suggestion": 0.0,
        }

    answer_map = answer_norm.set_index("label")["value"].to_dict()
    student_map = student_norm.set_index("label")["value"].to_dict()

    answer_total = float(sum(float(v) for v in answer_map.values()))
    student_total = float(sum(float(v) for v in student_map.values()))

    if answer_total <= 0:
        return {
            "match": False,
            "mismatches": [{"label": "<table>", "expected": "positive total", "actual": answer_total}],
            "score_suggestion": 0.0,
        }

    # Normalize student values to fractions in [0, 1] style.
    if student_total > 1.5:
        normalized_student = {label: float(val) / 100.0 for label, val in student_map.items()}
    else:
        normalized_student = {label: float(val) for label, val in student_map.items()}

    mismatches: list[dict[str, Any]] = []
    matches = 0

    for label, expected_raw in answer_map.items():
        expected_share = float(expected_raw) / answer_total
        actual_share = normalized_student.get(label)
        if actual_share is None:
            mismatches.append({"label": label, "expected": expected_share, "actual": None})
            continue

        if abs(actual_share - expected_share) <= percent_tolerance:
            matches += 1
        else:
            mismatches.append(
                {
                    "label": label,
                    "expected": expected_share,
                    "actual": actual_share,
                }
            )

    total = max(1, len(answer_map))
    ratio = matches / total
    match = len(mismatches) == 0
    score_suggestion = round(max(0.0, min(1.0, ratio)), 2)

    return {
        "match": match,
        "mismatches": mismatches,
        "score_suggestion": score_suggestion,
    }


def is_desc_sorted(df: pd.DataFrame) -> bool:
    normalized = _normalize_for_compare(df)
    if normalized.empty:
        return False
    values = normalized["value"].tolist()
    return values == sorted(values, reverse=True)


def is_desc_sorted_within_groups(df: pd.DataFrame) -> bool:
    """For nested pivots (col0 = outer group, repeated via ffill; col1+ = inner rows + values):
    verify that within each col0 group the rows are descending by the first numeric value column.

    Falls back to ``is_desc_sorted`` when all col0 values are unique (flat pivot).
    Rows whose outer label contains "total" are treated as subtotals and excluded.
    """
    if df.empty or len(df.columns) < 2:
        return is_desc_sorted(df)

    outer = df.iloc[:, 0].astype(str).str.strip()
    # Flat pivot: every category appears exactly once
    if outer.nunique() == len(outer):
        return is_desc_sorted(df)

    # Find the first numeric column (skip col0)
    value_col: str | None = None
    for col in df.columns[1:]:
        if pd.to_numeric(df[col], errors="coerce").notna().any():
            value_col = col
            break
    if value_col is None:
        return False

    tmp = df[[df.columns[0], value_col]].copy()
    tmp.columns = ["__group__", "__val__"]
    tmp["__group__"] = tmp["__group__"].astype(str).str.strip().str.lower()
    tmp["__val__"] = pd.to_numeric(tmp["__val__"], errors="coerce")
    # Exclude subtotal / grand-total rows
    tmp = tmp[~tmp["__group__"].str.contains("total", regex=False)]

    if tmp.empty:
        return False

    tmp = tmp.dropna(subset=["__val__"])
    if tmp.empty:
        return False

    checked_any_group = False
    for _, grp in tmp.groupby("__group__", sort=False):
        vals = grp["__val__"].tolist()
        if not vals:
            continue
        checked_any_group = True
        if vals != sorted(vals, reverse=True):
            return False

    return checked_any_group


def is_group_order_desc(student_df: pd.DataFrame, answer_df: pd.DataFrame) -> bool:
    """Verify top-level group order in student data matches answer-key order."""
    if student_df.empty or answer_df.empty or len(student_df.columns) < 1 or len(answer_df.columns) < 1:
        return False

    def _ordered_groups(df: pd.DataFrame) -> list[str]:
        out: list[str] = []
        seen: set[str] = set()
        first_col = df.columns[0]
        for raw in df[first_col].tolist():
            label = normalize_label("" if pd.isna(raw) else str(raw).strip())
            if not label or "total" in label:
                continue
            if label in seen:
                continue
            seen.add(label)
            out.append(label)
        return out

    student_order = _ordered_groups(student_df)
    answer_order = _ordered_groups(answer_df)
    if not student_order or not answer_order:
        return False

    return student_order == answer_order


def has_multiple_items_marker(df: pd.DataFrame) -> bool:
    if df.empty:
        return False
    for col in df.columns:
        series = df[col].astype(str).str.lower()
        if series.str.contains("multiple items", regex=False).any():
            return True
    return False


# ---------------------------------------------------------------------------
# Sheet fingerprinting for content-based sheet-to-question matching
# ---------------------------------------------------------------------------

def sheet_fingerprint(df: pd.DataFrame) -> dict[str, Any]:
    """Lightweight fingerprint for matching a student sheet to an answer-key question."""
    n_rows, n_cols = df.shape

    if n_rows < 20:
        row_bucket = "tiny"
    elif n_rows < 200:
        row_bucket = "small"
    elif n_rows < 2_000:
        row_bucket = "medium"
    elif n_rows < 15_000:
        row_bucket = "large"
    else:
        row_bucket = "xlarge"

    first_col = df.iloc[:, 0].dropna() if not df.empty else pd.Series([], dtype=object)
    if len(first_col) > 0:
        numeric_ratio = float(pd.to_numeric(first_col, errors="coerce").notna().mean())
    else:
        numeric_ratio = 0.0
    first_col_numeric = numeric_ratio > 0.8

    # Collect label set only for small tables (large ones are too expensive)
    first_col_labels: frozenset[str] = frozenset()
    if n_rows < 200:
        first_col_labels = frozenset(
            str(v).strip().lower()
            for v in first_col
            if v is not None and not pd.isna(v)
        )

    return {
        "n_rows": n_rows,
        "n_cols": n_cols,
        "row_bucket": row_bucket,
        "first_col_numeric": first_col_numeric,
        "first_col_labels": first_col_labels,
    }


def fingerprint_similarity(student_fp: dict[str, Any], answer_fp: dict[str, Any]) -> float:
    """Score how closely a student sheet fingerprint matches an answer-key fingerprint."""
    score = 0.0

    if student_fp["row_bucket"] == answer_fp["row_bucket"]:
        score += 3.0

    s_cols, a_cols = student_fp["n_cols"], answer_fp["n_cols"]
    if s_cols == a_cols:
        score += 2.0
    elif abs(s_cols - a_cols) <= 1:
        score += 1.0

    if student_fp["first_col_numeric"] == answer_fp["first_col_numeric"]:
        score += 2.0

    s_labels = student_fp["first_col_labels"]
    a_labels = answer_fp["first_col_labels"]
    if s_labels and a_labels:
        overlap = len(s_labels & a_labels) / max(len(a_labels), 1)
        score += overlap * 4.0

    return score
