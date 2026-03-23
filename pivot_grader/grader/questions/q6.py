from __future__ import annotations

from typing import Any

import pandas as pd
import openpyxl

from grader.pivot_checker import compare_pivot_values, evaluate_highlight_formatting
from grader.qualitative_grader import grade_explanation
from grader.utils.normalize import normalize_label


def _best_numeric_col(df: pd.DataFrame) -> str | None:
    fallback = None
    for col in df.columns[1:]:
        series = pd.to_numeric(df[col], errors="coerce").dropna()
        if series.empty:
            continue
        if series.max() <= 1.0:
            return str(col)
        if fallback is None:
            fallback = str(col)
    return fallback


def _extract_nested_vendor_product_values(df: pd.DataFrame, ws: Any | None = None) -> dict[str, float]:
    if df.empty or len(df.columns) < 2:
        return {}

    value_col = _best_numeric_col(df)
    if value_col is None:
        return {}

    out: dict[str, float] = {}
    current_vendor: str | None = None
    indent_by_row_offset: dict[int, bool] = {}

    if ws is not None:
        row_offset = 0
        for cell in ws["A"]:
            raw_value = "" if cell.value is None else str(cell.value)
            if not raw_value.strip():
                continue
            if row_offset == 0:
                # Skip the pivot header row (e.g. "Row Labels") to align with DataFrame rows.
                row_offset += 1
                continue
            alignment = cell.alignment
            indent = alignment.indent if alignment else 0
            indent_by_row_offset[row_offset - 1] = bool(indent and indent > 0)
            row_offset += 1

    for row_idx, row in df.iterrows():
        raw_label = "" if pd.isna(row.iloc[0]) else str(row.iloc[0])
        label = normalize_label(raw_label)
        if not label or "total" in label or label == "row labels":
            continue

        value = pd.to_numeric(row[value_col], errors="coerce")
        if pd.isna(value):
            continue

        if ws is not None:
            is_indented = indent_by_row_offset.get(int(row_idx), False)
        else:
            is_indented = raw_label != raw_label.lstrip()
        if not is_indented:
            current_vendor = label
            continue

        if current_vendor is None:
            continue

        out[f"{current_vendor}::{label}"] = float(value)

    return out


def _to_pct_of_vendor(df: pd.DataFrame, ws: Any | None = None) -> pd.DataFrame | None:
    if df.empty or len(df.columns) < 2:
        return None

    value_col = _best_numeric_col(df)
    if value_col is None:
        return None

    vendor_totals: dict[str, float] = {}
    product_values: list[tuple[str, str, float]] = []
    current_vendor: str | None = None
    indent_by_row_offset: dict[int, bool] = {}

    if ws is not None:
        row_offset = 0
        for cell in ws["A"]:
            raw_value = "" if cell.value is None else str(cell.value)
            if not raw_value.strip():
                continue
            if row_offset == 0:
                row_offset += 1
                continue
            alignment = cell.alignment
            indent = alignment.indent if alignment else 0
            indent_by_row_offset[row_offset - 1] = bool(indent and indent > 0)
            row_offset += 1

    for row_idx, row in df.iterrows():
        raw_label = "" if pd.isna(row.iloc[0]) else str(row.iloc[0])
        label = normalize_label(raw_label)
        if not label or "total" in label or label == "row labels":
            continue

        value = pd.to_numeric(row[value_col], errors="coerce")
        if pd.isna(value):
            continue

        if ws is not None:
            is_indented = indent_by_row_offset.get(int(row_idx), False)
        else:
            is_indented = raw_label != raw_label.lstrip()
        if not is_indented:
            current_vendor = label
            vendor_totals[current_vendor] = float(value)
            continue

        if current_vendor is None:
            continue

        product_values.append((current_vendor, label, float(value)))

    rows: list[dict[str, Any]] = []
    for vendor, product, raw_value in product_values:
        vendor_total = vendor_totals.get(vendor, 0.0)
        if vendor_total <= 0:
            continue
        rows.append({"Row Labels": product, "Value": raw_value / vendor_total})

    if not rows:
        return None
    return pd.DataFrame(rows)


def _extract_explanation_text(df: pd.DataFrame) -> str:
    text_values: list[str] = []
    for col in df.columns:
        series = df[col].dropna().astype(str)
        for value in series:
            value = value.strip()
            if len(value.split()) >= 6:
                text_values.append(value)
    return "\n".join(text_values)


def _filter_to_shared_product_rows(
    student_df: pd.DataFrame,
    answer_df: pd.DataFrame,
    numeric_tolerance: float = 1e-4,
) -> tuple[pd.DataFrame, pd.DataFrame] | None:
    if student_df.empty or answer_df.empty or len(student_df.columns) < 2 or len(answer_df.columns) < 2:
        return None

    student_value_col = _best_numeric_col(student_df)
    answer_value_col = _best_numeric_col(answer_df)
    if student_value_col is None or answer_value_col is None:
        return None

    student_label_col = student_df.columns[0]
    answer_label_col = answer_df.columns[0]

    student_tmp = student_df[[student_label_col, student_value_col]].copy()
    answer_tmp = answer_df[[answer_label_col, answer_value_col]].copy()
    student_tmp.columns = ["label", "value"]
    answer_tmp.columns = ["label", "value"]

    student_tmp["label"] = student_tmp["label"].apply(lambda x: normalize_label("" if pd.isna(x) else str(x)))
    answer_tmp["label"] = answer_tmp["label"].apply(lambda x: normalize_label("" if pd.isna(x) else str(x)))
    student_tmp["value"] = pd.to_numeric(student_tmp["value"], errors="coerce")
    answer_tmp["value"] = pd.to_numeric(answer_tmp["value"], errors="coerce")

    student_tmp = student_tmp.dropna(subset=["value"])
    answer_tmp = answer_tmp.dropna(subset=["value"])
    student_tmp = student_tmp[student_tmp["label"] != ""]
    answer_tmp = answer_tmp[answer_tmp["label"] != ""]
    student_tmp = student_tmp[~student_tmp["label"].str.contains("total")]
    answer_tmp = answer_tmp[~answer_tmp["label"].str.contains("total")]

    # Drop vendor subtotal artifacts that appear as ~1.0 in student outputs.
    student_tmp = student_tmp[
        ~(
            (student_tmp["value"] >= 1.0 - numeric_tolerance)
            & (student_tmp["value"] <= 1.0 + numeric_tolerance)
        )
    ]

    answer_labels = set(answer_tmp["label"].tolist())
    shared_labels = student_tmp[student_tmp["label"].isin(answer_labels)]["label"].tolist()
    if not shared_labels:
        return None

    shared_label_set = set(shared_labels)
    student_shared = student_tmp[student_tmp["label"].isin(shared_label_set)].copy()
    answer_shared = answer_tmp[answer_tmp["label"].isin(shared_label_set)].copy()
    if student_shared.empty or answer_shared.empty:
        return None

    return student_shared.rename(columns={"label": "Row Labels", "value": "Value"}), answer_shared.rename(
        columns={"label": "Row Labels", "value": "Value"}
    )


def grade_question(
    student_df: pd.DataFrame,
    answer_df: pd.DataFrame,
    question_cfg: dict[str, Any],
    workbook_path: Any = None,
    sheet_name: str | None = None,
    qid: str = "Q6",
) -> dict[str, Any]:
    value_issues: list[str] = []
    explanation_issues: list[str] = []
    worksheet = None
    workbook_handle = None

    if workbook_path and sheet_name:
        try:
            workbook_handle = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
            if sheet_name in workbook_handle.sheetnames:
                worksheet = workbook_handle[sheet_name]
        except Exception:
            worksheet = None
            workbook_handle = None

    try:
        if student_df.empty:
            formatting_score, formatting_issues = evaluate_highlight_formatting(workbook_path, sheet_name)
            return {
                "structural_score": 1.0,
                "value_score": 0.0,
                "formatting_score": formatting_score,
                "explanation_score": 1.0,
                "structural_issues": [],
                "value_issues": ["Missing pivot table"],
                "formatting_issues": formatting_issues,
                "explanation_issues": [],
            }

        value_result = compare_pivot_values(student_df, answer_df)
        used_raw_dollar_normalization = False

        if not value_result["match"]:
            shared = _filter_to_shared_product_rows(student_df, answer_df, numeric_tolerance=1e-4)
            if shared is not None:
                shared_student_df, shared_answer_df = shared
                shared_result = compare_pivot_values(
                    shared_student_df,
                    shared_answer_df,
                    numeric_tolerance=1e-4,
                )
                if shared_result["match"]:
                    value_result = shared_result

        if not value_result["match"]:
            student_vals = _extract_nested_vendor_product_values(student_df, ws=worksheet)
            if student_vals:
                max_student_val = max(student_vals.values())
                if max_student_val > 10.0:
                    normalized_df = _to_pct_of_vendor(student_df, ws=worksheet)
                    if normalized_df is not None:
                        value_result = compare_pivot_values(normalized_df, answer_df)
                        used_raw_dollar_normalization = value_result["match"]

        value_score = 1.0 if value_result["match"] else 0.0
        if used_raw_dollar_normalization:
            value_issues.append("Values shown as raw dollars, not % of parent row")
        elif not value_result["match"] and value_result["mismatches"]:
            first = value_result["mismatches"][0]
            value_issues.append(
                f"Mismatch at {first['label']}: expected {first['expected']}, actual {first['actual']}."
            )

        explanation_score = 1.0
        if question_cfg.get("explanation_required"):
            rubric_text = question_cfg.get("explanation_rubric", "")
            student_text = _extract_explanation_text(student_df)
            llm_result = grade_explanation(qid, student_text, rubric_text)
            if llm_result.get("needs_review", False):
                explanation_score = 0.0
                reason = str(llm_result.get("brief_reason", "Manual review needed"))
                explanation_issues.append(f"NEEDS_REVIEW: {reason}")
            elif llm_result.get("deduct_explanation", False):
                explanation_score = 0.0
                explanation_issues.append(str(llm_result.get("brief_reason", "Needs more detail")))

        formatting_score, formatting_issues = evaluate_highlight_formatting(workbook_path, sheet_name)

        return {
            "structural_score": 1.0,
            "value_score": value_score,
            "formatting_score": formatting_score,
            "explanation_score": explanation_score,
            "structural_issues": [],
            "value_issues": value_issues,
            "formatting_issues": formatting_issues,
            "explanation_issues": explanation_issues,
        }
    finally:
        if workbook_handle is not None:
            workbook_handle.close()