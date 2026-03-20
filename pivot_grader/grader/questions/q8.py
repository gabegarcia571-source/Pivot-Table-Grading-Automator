from __future__ import annotations

from typing import Any

import pandas as pd


def _cell_is_highlighted(cell: Any) -> bool:
    fill = cell.fill
    if fill is None or fill.fill_type in (None, "none"):
        return False
    fg = fill.fgColor
    if fg is None:
        return False
    if fg.type in ("theme", "indexed"):
        return False
    rgb = (fg.rgb if hasattr(fg, "rgb") else "").upper()
    return rgb not in ("", "00000000", "FFFFFFFF", "FF000000")


def check_q8_highlight(workbook_path: Any, sheet_name: str) -> dict[str, Any]:
    from pathlib import Path as _Path

    import openpyxl as _openpyxl

    from grader.answer_constants import HOLIDAY_ONLY_CUSTOMERS, HOLIDAY_ONLY_CUSTOMERS_COMPLETE

    notes: list[str] = []

    if not HOLIDAY_ONLY_CUSTOMERS_COMPLETE:
        return {
            "match": False,
            "missing_pivot": False,
            "needs_review": True,
            "notes": ["NEEDS_REVIEW: Q8 answer set is incomplete"],
        }

    try:
        wb = _openpyxl.load_workbook(_Path(workbook_path), data_only=True)
    except Exception:
        notes.append("Missing highlight")
        return {"match": False, "missing_pivot": False, "needs_review": False, "notes": notes}

    if sheet_name not in wb.sheetnames:
        notes.append("Missing highlight")
        wb.close()
        return {"match": False, "missing_pivot": False, "needs_review": False, "notes": notes}

    ws = wb[sheet_name]
    highlighted_ids: set[int] = set()

    for row in ws.iter_rows():
        if not row:
            continue
        row_highlighted = any(_cell_is_highlighted(cell) for cell in row)
        if not row_highlighted:
            continue
        label_val = row[0].value
        try:
            highlighted_ids.add(int(label_val))
        except (ValueError, TypeError):
            pass

    wb.close()

    if not highlighted_ids:
        notes.append("Missing highlight")
        return {"match": False, "missing_pivot": True, "needs_review": False, "notes": notes}

    correct_set = HOLIDAY_ONLY_CUSTOMERS
    n_correct = max(1, len(correct_set))
    false_positives = highlighted_ids - correct_set
    false_negatives = correct_set - highlighted_ids
    error_count = len(false_positives) + len(false_negatives)
    error_rate = error_count / n_correct

    if error_rate == 0.0:
        return {"match": True, "missing_pivot": False, "needs_review": False, "notes": notes}

    if error_rate <= 0.05:
        return {"match": True, "missing_pivot": False, "needs_review": False, "notes": notes}

    notes.append("Incorrect value")
    return {"match": False, "missing_pivot": False, "needs_review": False, "notes": notes}


def grade_question(
    student_df: pd.DataFrame,
    answer_df: pd.DataFrame,
    question_cfg: dict[str, Any],
    workbook_path: Any = None,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    if workbook_path and sheet_name:
        q8 = check_q8_highlight(workbook_path, sheet_name)
    else:
        q8 = {
            "match": False,
            "missing_pivot": False,
            "needs_review": True,
            "notes": ["NEEDS_REVIEW: workbook path unavailable for Q8 highlight detection"],
        }

    value_score = 1.0 if q8["match"] else 0.0
    value_issues = list(q8.get("notes", []))

    return {
        "structural_score": 1.0,
        "value_score": value_score,
        "formatting_score": 1.0,
        "explanation_score": 1.0,
        "structural_issues": [],
        "value_issues": value_issues,
        "formatting_issues": [],
        "explanation_issues": [],
    }
