from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from grader.grade_writer import write_grades
from grader.ingest import find_pivot_origin, load_student_submission
from grader.pivot_checker import (
    compare_pivot_values,
    compare_pivot_values_subset,
    evaluate_highlight_formatting,
    has_any_highlight,
    is_group_order_desc,
    is_desc_sorted_within_groups,
    sheet_fingerprint,
    fingerprint_similarity,
)
from grader.questions.q3 import evaluate_q3_structure
from grader.questions.q4 import check_q4_average
from grader.questions.q5 import check_q5_filter, grade_question as grade_q5
from grader.questions.q6 import (
    _best_numeric_col,
    _extract_nested_vendor_product_values,
    grade_question as grade_q6,
)
from grader.questions.q7 import _compare_maps
from grader.questions.q8 import grade_question as grade_q8
from grader.questions.q9 import grade_question as grade_q9
from grader.questions.q10 import check_q10_filter, grade_question as grade_q10


def test_find_pivot_origin_standard() -> None:
    """Pivot at A1 — origin should be (1, 1)."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Label"
    ws["B1"] = "Value"
    ws["A2"] = "Food & Drink"
    ws["B2"] = 1_707_732.18

    assert find_pivot_origin(ws) == (1, 1)


def test_find_pivot_origin_shifted() -> None:
    """Stray label in row 3, actual pivot header starting at C8."""
    wb = Workbook()
    ws = wb.active
    ws["A3"] = "Student Name"   # single-cell stray label — only 1 non-empty in row
    ws["C8"] = "category_name"
    ws["D8"] = "Sum of total_product_price"
    ws["C9"] = "Food & Drink"
    ws["D9"] = 1_707_732.18

    assert find_pivot_origin(ws) == (8, 3)


def test_load_student_submission_shifted_pivot(tmp_path: Path) -> None:
    """Ensure the pivot is found and loaded correctly when offset from A1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1"
    ws["A3"] = "My Homework"              # stray single-cell label
    ws["C8"] = "category_name"            # pivot header starts here
    ws["D8"] = "Sum of total_product_price"
    ws["C9"] = "Food & Drink"
    ws["D9"] = 1_707_732.18

    student_dir = tmp_path / "student_1"
    student_dir.mkdir()
    wb.save(student_dir / "submission.xlsx")

    result = load_student_submission(student_dir)

    assert result.error is None
    assert "Q1" in result.sheets
    df = result.sheets["Q1"]
    # First column must be the pivot label column, not an offset blank column
    assert df.columns[0] == "category_name"
    assert len(df) >= 1


def test_load_student_submission_from_file_path(tmp_path: Path) -> None:
    """load_student_submission should also accept a direct xlsx path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1"
    ws["A1"] = "category_name"
    ws["B1"] = "Sum of total_product_price"
    ws["A2"] = "Food & Drink"
    ws["B2"] = 1_707_732.18

    workbook_path = tmp_path / "student_file.xlsx"
    wb.save(workbook_path)

    result = load_student_submission(workbook_path)

    assert result.error is None
    assert result.student_id == "student_file"
    assert result.workbook_path == workbook_path
    assert "Q1" in result.sheets


# ---------------------------------------------------------------------------
# Q4 average-scan tests
# ---------------------------------------------------------------------------

def test_check_q4_average_match() -> None:
    """A value within 2.0 of 149.47 should produce match=True."""
    df = pd.DataFrame({"order_id": [1, 2, 3], "order_total": [144.5, 149.48, 150.0]})
    result = check_q4_average(df)
    assert result["has_numeric"] is True
    assert result["match"] is True
    assert abs(result["closest"] - 149.47) <= 2.0


def test_check_q4_average_wrong_values() -> None:
    """Values that exist but none near 149.47 — wrong_values deduction expected."""
    df = pd.DataFrame({"order_id": [1, 2, 3], "line_avg": [12.50, 13.00, 11.75]})
    result = check_q4_average(df)
    assert result["has_numeric"] is True
    assert result["match"] is False
    assert result["delta"] > 2.0


def test_check_q4_average_no_numerics() -> None:
    """A sheet with no numeric values at all — missing_pivot deduction expected."""
    df = pd.DataFrame({"note": ["See other sheet", "N/A"], "col2": [None, None]})
    result = check_q4_average(df)
    assert result["has_numeric"] is False
    assert result["match"] is False
    assert result["closest"] is None


def test_compare_pivot_values_match() -> None:
    student = pd.DataFrame({"Label": ["A", "B"], "Value": [10.0, 20.0]})
    answer = pd.DataFrame({"Label": ["A", "B"], "Value": [10.0, 20.0]})

    result = compare_pivot_values(student, answer)

    assert result["match"] is True
    assert result["mismatches"] == []
    assert result["score_suggestion"] == 1.0


def test_compare_pivot_values_partial_credit() -> None:
    student = pd.DataFrame({"Label": ["A", "B"], "Value": [10.0, 22.0]})
    answer = pd.DataFrame({"Label": ["A", "B"], "Value": [10.0, 20.0]})

    result = compare_pivot_values(student, answer)

    assert result["match"] is False
    assert len(result["mismatches"]) == 1
    assert result["score_suggestion"] == 0.5


def test_compare_pivot_values_subset_allows_additional_filtering() -> None:
    """Q3-style filtered subset should pass when required labels are present and correct."""
    student = pd.DataFrame({
        "Row Labels": ["Sweetums Soda", "Sweetums Candy", "Zoo Penguin Plush"],
        "Value": [5604, 5312, 5279],
    })
    answer = pd.DataFrame({
        "Row Labels": ["Sweetums Soda", "Sweetums Candy", "Zoo Penguin Plush", "Waffle Ale"],
        "Value": [5604, 5312, 5279, 5206],
    })

    result = compare_pivot_values_subset(
        student,
        answer,
        required_labels=["Sweetums Soda", "Sweetums Candy", "Zoo Penguin Plush"],
    )

    assert result["match"] is True
    assert result["mismatches"] == []
    assert result["missing_required"] == []


def test_compare_pivot_values_subset_requires_key_labels() -> None:
    """Subset match should fail if one required top product is missing."""
    student = pd.DataFrame({
        "Row Labels": ["Sweetums Soda", "Sweetums Candy"],
        "Value": [5604, 5312],
    })
    answer = pd.DataFrame({
        "Row Labels": ["Sweetums Soda", "Sweetums Candy", "Zoo Penguin Plush", "Waffle Ale"],
        "Value": [5604, 5312, 5279, 5206],
    })

    result = compare_pivot_values_subset(
        student,
        answer,
        required_labels=["Sweetums Soda", "Sweetums Candy", "Zoo Penguin Plush"],
    )

    assert result["match"] is False
    assert result["missing_required"] == ["Zoo Penguin Plush"]


def test_compare_pivot_values_subset_can_ignore_q3_vendor_group_headers() -> None:
    """Q3 value matching should ignore vendor group header rows when requested."""
    student = pd.DataFrame({
        "Row Labels": [
            "Sweetums Industries",
            "Sweetums Soda",
            "Sweetums Candy",
            "Lil' Sebastian Co.",
            "Zoo Penguin Plush",
        ],
        "Value": [13566, 5604, 5312, 5279, 5279],
    })
    answer = pd.DataFrame({
        "Row Labels": ["Sweetums Soda", "Sweetums Candy", "Zoo Penguin Plush"],
        "Value": [5604, 5312, 5279],
    })

    result = compare_pivot_values_subset(
        student,
        answer,
        required_labels=["Sweetums Soda", "Sweetums Candy", "Zoo Penguin Plush"],
        ignore_labels=["Sweetums Industries", "Lil' Sebastian Co."],
    )

    assert result["match"] is True
    assert result["mismatches"] == []


def test_q6_best_numeric_col_prefers_percent_column() -> None:
    df = pd.DataFrame(
        {
            "Row Labels": ["A", "B"],
            "Raw Dollars": [1200.0, 800.0],
            "% of Vendor": [0.6, 0.4],
        }
    )

    assert _best_numeric_col(df) == "% of Vendor"


def test_q6_extract_nested_values_uses_alignment_indent() -> None:
    df = pd.DataFrame(
        {
            "Row Labels": ["Sweetums", "Soda", "Candy"],
            "Value": [1000.0, 0.6, 0.4],
        }
    )

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Row Labels"
    ws["A2"] = "Sweetums"
    ws["A3"] = "Soda"
    ws["A3"].alignment = Alignment(indent=1)
    ws["A4"] = "Candy"
    ws["A4"].alignment = Alignment(indent=1)

    nested = _extract_nested_vendor_product_values(df, ws=ws)

    assert nested == {
        "sweetums::soda": 0.6,
        "sweetums::candy": 0.4,
    }


def test_q6_partial_expansion_shared_product_labels_can_pass() -> None:
    student = pd.DataFrame(
        {
            "Row Labels": ["Sweetums", "Soda", "Candy"],
            "Value": [1.0, 0.6, 0.4],
        }
    )
    answer = pd.DataFrame(
        {
            "Row Labels": ["Soda", "Candy", "Waffle Ale", "Pretzel"],
            "Value": [0.6, 0.4, 0.7, 0.3],
        }
    )

    result = grade_q6(student, answer, question_cfg={"explanation_required": False})

    assert result["value_score"] == 1.0
    assert result["value_issues"] == []


def test_q6_partial_expansion_fails_when_shared_label_value_wrong() -> None:
    student = pd.DataFrame(
        {
            "Row Labels": ["Sweetums", "Soda", "Candy"],
            "Value": [1.0, 0.61, 0.4],
        }
    )
    answer = pd.DataFrame(
        {
            "Row Labels": ["Soda", "Candy", "Waffle Ale", "Pretzel"],
            "Value": [0.6, 0.4, 0.7, 0.3],
        }
    )

    result = grade_q6(student, answer, question_cfg={"explanation_required": False})

    assert result["value_score"] == 0.0
    assert result["value_issues"]


def test_write_grades(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    output_dir = tmp_path / "out"

    wb = Workbook()
    ws = wb.active
    ws.title = "Gradesheet"
    ws["C4"] = "=C17"
    ws["C17"] = "=SUM(C7:C16)"
    wb.save(template)

    scores = {f"Q{i}": 1.0 for i in range(1, 11)}
    comments = {f"Q{i}": "ok" for i in range(1, 11)}

    out_file = write_grades(
        student_id="student_1",
        scores_dict=scores,
        comments_dict=comments,
        template_path=template,
        output_dir=output_dir,
    )

    graded = load_workbook(out_file)
    gws = graded["Gradesheet"]

    assert gws["C3"].value == "student_1"
    assert gws["C7"].value == 1.0
    assert gws["D7"].value == "ok"
    assert gws["C4"].value == "=C17"
    assert gws["C17"].value == "=SUM(C7:C16)"


# ---------------------------------------------------------------------------
# ffill merged-cell tests
# ---------------------------------------------------------------------------

def test_load_student_submission_merged_cell_ffill(tmp_path: Path) -> None:
    """Merged pivot row-group labels (NaN after export) must be filled down."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1"
    # Simulate a pivot with a row-group label merged across 3 rows:
    # Excel exports the label only in the first row; remaining cells are blank.
    ws["A1"] = "Category"
    ws["B1"] = "Value"
    ws["A2"] = "Food & Drink"  # exported label (first row of group)
    ws["B2"] = 100.0
    ws["A3"] = None            # merged — blank in export
    ws["B3"] = 200.0
    ws["A4"] = None            # merged — blank in export
    ws["B4"] = 150.0

    student_dir = tmp_path / "student_merged"
    student_dir.mkdir()
    wb.save(student_dir / "submission.xlsx")

    result = load_student_submission(student_dir)
    assert result.error is None
    df = result.sheets["Q1"]

    first_col = df.iloc[:, 0].tolist()
    # After ffill every cell should hold "Food & Drink", not NaN
    assert all(v == "Food & Drink" for v in first_col), (
        f"Expected ffill to propagate 'Food & Drink'; got {first_col}"
    )


# ---------------------------------------------------------------------------
# Q3 filter-check tests
# ---------------------------------------------------------------------------

_VENDOR_NAMES = [
    "sweetums industries",
    "jj's diner goods",
    "rent-a-swag inc.",
    "lil' sebastian co.",
]


def test_check_q3_filter_not_applied() -> None:
    """All four vendor names in first column → filter_ok=False."""
    df = pd.DataFrame({
        "vendor_name": _VENDOR_NAMES,
        "total": [1000, 2000, 1500, 800],
    })
    result = evaluate_q3_structure(df)
    assert result["structure_ok"] is False
    assert len(result["found_vendors_in_rows"]) == 4


def test_check_q3_filter_applied() -> None:
    """Product names in first column, no vendor names → filter_ok=True."""
    df = pd.DataFrame({
        "product_name": ["Calzone", "Pizza", "Waffle"],
        "total": [500, 600, 700],
    })
    result = evaluate_q3_structure(df)
    assert result["structure_ok"] is True
    assert result["found_vendors_in_rows"] == []


def test_check_q3_filter_partial_vendors() -> None:
    """Any vendor names in first column mean vendor split rows → filter_ok=False."""
    df = pd.DataFrame({
        "vendor_name": ["sweetums industries", "jj's diner goods"],
        "total": [1000, 2000],
    })
    result = evaluate_q3_structure(df)
    assert result["structure_ok"] is False
    assert set(result["found_vendors_in_rows"]) == {"sweetums industries", "jj's diner goods"}


def test_check_q3_filter_correct_top3() -> None:
    """Top-3 vendor row-label style is structurally invalid for Q3."""
    df = pd.DataFrame({
        "vendor_name": ["sweetums industries", "jj's diner goods", "lil' sebastian co."],
        "total": [1000, 2000, 800],
    })
    result = evaluate_q3_structure(df)
    assert result["structure_ok"] is False
    assert result["extra_vendors"] == []
    assert result["missing_vendors"] == []


def test_evaluate_q3_structure_reports_style_and_marker_state() -> None:
    df = pd.DataFrame({
        "Row Labels": ["Sweetums Industries", "JJ's Diner Goods", "Lil' Sebastian Co."],
        "Value": [1000, 2000, 800],
    })
    result = evaluate_q3_structure(df)
    assert result["structure_ok"] is False
    assert result["style"] == "row_label_vendor"
    assert set(result["found_vendors_in_rows"]) == {
        "sweetums industries",
        "jj's diner goods",
        "lil' sebastian co.",
    }


def test_check_q3_filter_handles_smart_apostrophes() -> None:
    """Curly apostrophes should still normalize for detecting invalid vendor rows."""
    df = pd.DataFrame({
        "vendor_name": ["JJ’s Diner Goods", "Lil’ Sebastian Co.", "Sweetums Industries"],
        "total": [1000, 800, 2000],
    })
    result = evaluate_q3_structure(df)
    assert result["structure_ok"] is False
    assert result["missing_vendors"] == []


# ---------------------------------------------------------------------------
# Q5 filter-check tests
# ---------------------------------------------------------------------------

def test_check_q5_filter_correct() -> None:
    """Only Jun/Jul/Aug labels → filter_ok=True."""
    df = pd.DataFrame({
        "month": ["Jun", "Jul", "Aug"],
        "total": [300, 400, 350],
    })
    result = check_q5_filter(df)
    assert result["filter_ok"] is True
    assert result["non_summer_months_found"] == []
    assert result["summer_months_found"]


def test_check_q5_filter_wrong() -> None:
    """January label present → filter_ok=False."""
    df = pd.DataFrame({
        "month": ["Jan", "Jun", "Jul", "Aug"],
        "total": [200, 300, 400, 350],
    })
    result = check_q5_filter(df)
    assert result["filter_ok"] is False
    assert "jan" in result["non_summer_months_found"]


def test_check_q5_filter_fails_when_no_summer_months_present() -> None:
    """(All)-style pivot with no month breakdown should fail the filter check."""
    df = pd.DataFrame(
        {
            "Row Labels": ["Food & Drink", "Home & Lifestyle", "Clothing & Accessories"],
            "order_short_date": ["(All)", "(All)", "(All)"],
            "Value": [100.0, 80.0, 60.0],
        }
    )

    result = check_q5_filter(df)
    assert result["filter_ok"] is False
    assert result["non_summer_months_found"] == []
    assert result["summer_months_found"] == []


def test_grade_q5_reports_no_summer_filter_not_applied_message() -> None:
    student_df = pd.DataFrame(
        {
            "Row Labels": ["Food & Drink", "Clothing & Accessories"],
            "order_short_date": ["(All)", "(All)"],
            "Sum of total_product_price": [16160, 9746],
        }
    )
    answer_df = pd.DataFrame(
        {
            "Row Labels": ["Food & Drink", "Jun", "Jul", "Aug"],
            "Sum of total_product_price": [16160, 5858, 5584, 4718],
        }
    )

    result = grade_q5(student_df, answer_df, question_cfg={})

    assert result["structural_score"] == 0.0
    assert "Month filter not applied - no summer months found in pivot" in result["structural_issues"]


def test_grade_q5_non_summer_message_only_when_summer_present() -> None:
    student_df = pd.DataFrame(
        {
            "Row Labels": ["Jun", "Jul", "Jan"],
            "category_name": ["Food & Drink", "Clothing & Accessories", "Home & Lifestyle"],
            "Sum of total_product_price": [100.0, 80.0, 60.0],
        }
    )
    answer_df = pd.DataFrame(
        {
            "Row Labels": ["Food & Drink", "Jun", "Jul", "Aug"],
            "Sum of total_product_price": [100.0, 40.0, 35.0, 25.0],
        }
    )

    result = grade_q5(student_df, answer_df, question_cfg={})
    assert result["structural_score"] == 0.0
    assert any("Month filter incorrect - non-summer months found:" in s for s in result["structural_issues"])
    assert all("Month filter not applied - no summer months found in pivot" not in s for s in result["structural_issues"])


def test_grade_q5_ignores_month_subrows_in_answer_values() -> None:
    """Q5 value check should compare category totals, not nested month detail rows."""
    student_df = pd.DataFrame({
        "Row Labels": ["Food & Drink", "Clothing & Accessories"],
        "order_short_date": ["Jun", "Jul"],
        "Sum of total_product_price": [16160, 9746],
    })
    answer_df = pd.DataFrame({
        "Row Labels": [
            "Food & Drink",
            "Jun",
            "Jul",
            "Aug",
            "Clothing & Accessories",
        ],
        "Sum of total_product_price": [16160, 5858, 5584, 4718, 9746],
    })

    result = grade_q5(student_df, answer_df, question_cfg={})

    assert result["structural_score"] == 1.0
    assert result["value_score"] == 1.0
    assert result["value_issues"] == []


# ---------------------------------------------------------------------------
# Q10 structure-check tests (vendor count, column headers, proportion values)
# ---------------------------------------------------------------------------

_Q10_VENDORS = [
    "Sweetums Industries",
    "KnopeWorks",
    "JJ's Diner Goods",
    "Perd Products",
    "Rent-A-Swag Inc.",
    "Lil' Sebastian Co.",
    "Pyramid Outfitters",
    "Pawnee Goddesses Collective",
    "Entertainment 720",
    "Treat Yo' Self LLC",
    "Mouse Rat Merchandising",
    "Eagleton Elegance",
    "Snakehole Spirits Co.",
    "Burt Macklin Enterprises",
]


def test_check_q10_filter_correct() -> None:
    """14 vendor rows with honey/no-promo data and numerics should pass filter check."""
    df = pd.DataFrame({
        "vendor_name": _Q10_VENDORS,
        "Honey": [0.07, 0.07, 0.06, 0.07, 0.20, 0.05, 0.06, 0.07, 0.06, 0.08, 0.06, 0.06, 0.09, 0.05],
        "No Promo Code": [0.93, 0.93, 0.94, 0.93, 0.80, 0.95, 0.94, 0.93, 0.94, 0.92, 0.94, 0.94, 0.91, 0.95],
        "Grand Total": [1.00] * 14,
    })
    result = check_q10_filter(df)
    assert result["filter_ok"] is True
    assert result["vendor_count_ok"] is True
    assert result["has_honey_col"] is True
    assert result["has_no_promo_col"] is True
    assert result["numeric_present"] is True


def test_check_q10_filter_wrong_vendor_count() -> None:
    """Only 4 vendor rows → filter_ok=False, vendor_count_ok=False."""
    df = pd.DataFrame({
        "vendor_name": _Q10_VENDORS[:4],
        "Honey": [0.21, 0.15, 0.53, 0.40],
        "No Promo Code": [0.79, 0.85, 0.47, 0.60],
    })
    result = check_q10_filter(df)
    assert result["filter_ok"] is False
    assert result["vendor_count_ok"] is False
    assert result["vendor_count"] == 4


def test_check_q10_filter_missing_honey_column() -> None:
    """No 'Honey' column header → filter_ok=False, has_honey_col=False."""
    df = pd.DataFrame({
        "vendor_name": _Q10_VENDORS,
        "No Promo Code": [0.79] * 14,
        "Grand Total": [1.00] * 14,
    })
    result = check_q10_filter(df)
    assert result["filter_ok"] is False
    assert result["has_honey_col"] is False


def test_check_q10_filter_values_not_proportions() -> None:
    """Raw counts should fail value-range check for proportion-based Q10 output."""
    df = pd.DataFrame({
        "vendor_name": _Q10_VENDORS,
        "Honey": [210, 150, 530, 400, 300, 100, 250, 180, 190, 260, 140, 110, 170, 120],
        "No Promo Code": [790, 850, 470, 600, 700, 900, 750, 820, 810, 740, 860, 890, 830, 880],
    })
    result = check_q10_filter(df)
    assert result["filter_ok"] is False
    assert result["numeric_present"] is True
    assert result["values_in_range"] is False


# ---------------------------------------------------------------------------
# Q1 within-group sort tests
# ---------------------------------------------------------------------------

def test_is_desc_sorted_within_groups_correct() -> None:
    """Days within each category are descending — should pass."""
    df = pd.DataFrame({
        "category": ["Food & Drink", "Food & Drink", "Food & Drink",
                     "Electronics",  "Electronics",  "Electronics"],
        "day":      ["Monday", "Tuesday", "Wednesday", "Friday", "Saturday", "Sunday"],
        "value":    [300.0, 200.0, 100.0, 150.0, 120.0, 80.0],
    })
    assert is_desc_sorted_within_groups(df) is True


def test_is_desc_sorted_within_groups_wrong_order() -> None:
    """One group has ascending order — should fail."""
    df = pd.DataFrame({
        "category": ["Food & Drink", "Food & Drink", "Electronics", "Electronics"],
        "day":      ["Monday", "Tuesday", "Friday", "Saturday"],
        "value":    [100.0, 200.0, 150.0, 80.0],  # Food & Drink ascending — wrong
    })
    assert is_desc_sorted_within_groups(df) is False


def test_is_desc_sorted_within_groups_flat_fallback() -> None:
    """No repeated first-col labels — falls back to flat is_desc_sorted, passes."""
    df = pd.DataFrame({
        "category": ["Electronics", "Food & Drink", "Outdoors"],
        "value":    [300.0, 200.0, 100.0],
    })
    assert is_desc_sorted_within_groups(df) is True


def test_is_desc_sorted_within_groups_ignores_total_rows() -> None:
    """Subtotal row (label contains 'total') is excluded from the sort check."""
    df = pd.DataFrame({
        "category": ["Food & Drink", "Food & Drink", "Food & Drink Total",
                     "Electronics",  "Electronics",  "Electronics Total"],
        "day":      ["Monday", "Tuesday", "Total", "Friday", "Saturday", "Total"],
        "value":    [300.0, 200.0, 500.0, 150.0, 80.0, 230.0],  # totals would break sort
    })
    assert is_desc_sorted_within_groups(df) is True


# ---------------------------------------------------------------------------
# Q8 highlight-detection tests (shared has_any_highlight utility)
# ---------------------------------------------------------------------------

from openpyxl.styles import Color, PatternFill

_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _make_highlight_workbook(
    tmp_path: Path,
    highlighted_rows: list[int],
    use_theme_fill: bool = False,
) -> tuple[Path, str]:
    """Write a minimal xlsx and highlight selected rows. Returns (path, sheet_name)."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Q8"
    ws["A1"] = "customer_id"
    ws["B1"] = "Nov"
    ws["C1"] = "Dec"
    for row, cid in enumerate([1001, 1002, 1003, 1004, 1005], start=2):
        ws.cell(row=row, column=1, value=cid)
        if row in highlighted_rows:
            if use_theme_fill:
                ws.cell(row=row, column=1).fill = PatternFill(fill_type="solid", fgColor=Color(theme=1))
            else:
                ws.cell(row=row, column=1).fill = _YELLOW_FILL
    path = tmp_path / "q8_test.xlsx"
    wb.save(path)
    return path, "Q8"


def test_has_any_highlight_true_on_solid_fill(tmp_path: Path) -> None:
    path, sname = _make_highlight_workbook(tmp_path, highlighted_rows=[3])
    assert has_any_highlight(path, sname) is True


def test_has_any_highlight_false_when_none(tmp_path: Path) -> None:
    path, sname = _make_highlight_workbook(tmp_path, highlighted_rows=[])
    assert has_any_highlight(path, sname) is False


def test_has_any_highlight_true_on_theme_fill(tmp_path: Path) -> None:
    path, sname = _make_highlight_workbook(tmp_path, highlighted_rows=[4], use_theme_fill=True)
    assert has_any_highlight(path, sname) is True


def test_has_any_highlight_false_on_missing_sheet(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "WrongSheet"
    path = tmp_path / "q8_wrong_sheet.xlsx"
    wb.save(path)
    assert has_any_highlight(path, "Q8") is False


# ---------------------------------------------------------------------------
# sheet_fingerprint / fingerprint_similarity tests
# ---------------------------------------------------------------------------

def test_sheet_fingerprint_row_bucket() -> None:
    tiny_df = pd.DataFrame({"a": range(5), "b": range(5)})
    medium_df = pd.DataFrame({"a": range(500), "b": range(500)})
    large_df = pd.DataFrame({"a": range(3000), "b": range(3000)})
    assert sheet_fingerprint(tiny_df)["row_bucket"] == "tiny"
    assert sheet_fingerprint(medium_df)["row_bucket"] == "medium"
    assert sheet_fingerprint(large_df)["row_bucket"] == "large"


def test_fingerprint_similarity_identical() -> None:
    df = pd.DataFrame({
        "label": ["Food & Drink", "Electronics"],
        "total": [1000.0, 2000.0],
    })
    fp = sheet_fingerprint(df)
    # Identical fingerprints should yield the maximum score
    score = fingerprint_similarity(fp, fp)
    assert score >= 9.0  # row_bucket(3) + cols_exact(2) + first_col_numeric(2) + label_overlap(4×1.0)


def test_fingerprint_similarity_different_sizes() -> None:
    small_df = pd.DataFrame({"label": ["A", "B"], "total": [1.0, 2.0]})
    large_df = pd.DataFrame({"x": range(5000), "y": range(5000), "z": range(5000)})
    fp_small = sheet_fingerprint(small_df)
    fp_large = sheet_fingerprint(large_df)
    low_score = fingerprint_similarity(fp_small, fp_large)
    # Different row bucket (tiny vs large) → no row_bucket bonus; column mismatch → low score
    assert low_score < 3.0


def test_is_group_order_desc_fails_wrong_top_category_order() -> None:
    answer_df = pd.DataFrame(
        {
            "Category": [
                "Food & Drink",
                "Food & Drink",
                "Clothing & Accessories",
                "Clothing & Accessories",
            ],
            "Value": [275.0, 225.0, 250.0, 200.0],
        }
    )
    df = pd.DataFrame(
        {
            "Category": [
                "Clothing & Accessories",
                "Clothing & Accessories",
                "Food & Drink",
                "Food & Drink",
            ],
            "Value": [250.0, 200.0, 275.0, 225.0],
        }
    )

    assert is_desc_sorted_within_groups(df) is True
    assert is_group_order_desc(df, answer_df) is False


def test_q7_compare_maps_detects_zero_value_mismatch() -> None:
    answer_map = {
        "food drink::kids teens": 0.0,
        "food drink::adults": 0.9,
    }
    student_map = {
        "food drink::kids teens": 0.2,
        "food drink::adults": 0.9,
    }

    result = _compare_maps(student_map, answer_map)

    assert result["match"] is False
    assert any(m["label"] == "food drink::kids teens" for m in result["mismatches"])


def test_q8_grade_question_deducts_wrong_measure() -> None:
    student_df = pd.DataFrame(
        {
            "customer_id": list(range(1, 1001)),
            "Nov": [1] * 1000,
            "Dec": [1] * 1000,
            "Count of order_id": [2] * 1000,
        }
    )

    result = grade_q8(student_df, pd.DataFrame(), question_cfg={})

    assert result["structural_score"] == 1.0
    assert result["value_score"] == 0.0
    assert "Wrong measure: used Count instead of Sum of total_product_price" in result["value_issues"]


def test_q9_grade_question_reports_measure_mismatch() -> None:
    answer_df = pd.DataFrame(
        {
            "Row Labels": [
                "4",
                "Honey",
                "No Promo Code",
                "12",
                "Honey",
                "No Promo Code",
                "27",
                "Honey",
                "No Promo Code",
            ],
            "Count of promo_code": [2, 1, 1, 2, 1, 1, 2, 1, 1],
        }
    )
    student_df = pd.DataFrame(
        [
            ["Row Labels", "Honey", "No Promo Code", "Grand Total", "Count of order_id"],
            [4, 1, 1, 2, None],
            [12, 1, 1, 2, None],
            [27, 1, 1, 2, None],
        ]
    )

    result = grade_q9(student_df, answer_df, question_cfg={"explanation_required": False})

    assert result["value_score"] == 0.0
    assert "Wrong measure: used Count of order_id instead of Count of promo_code" in result["value_issues"]


def test_q9_grade_question_flags_alternate_matrix_layout_for_review() -> None:
    answer_df = pd.DataFrame(
        {
            "Row Labels": ["4", "12", "27"],
            "Count of promo_code": [2, 2, 2],
        }
    )
    student_df = pd.DataFrame(
        [
            ["Row Labels", "Honey", "No Promo Code", "Count of promo_code"],
            [4, 1, 1, None],
            [12, 1, 1, None],
            [27, 1, 1, None],
        ]
    )

    result = grade_q9(student_df, answer_df, question_cfg={"explanation_required": False})

    assert any("NEEDS_REVIEW: alternate layout detected" in s for s in result["structural_issues"])
    assert result["needs_review"] is True
    assert result["value_score"] == 0.0


def test_q10_grade_question_can_return_full_credit() -> None:
    vendors = _Q10_VENDORS
    student_df = pd.DataFrame(
        {
            "vendor_name": vendors,
            "Honey": [0.05] * len(vendors),
            "No Promo Code": [0.95] * len(vendors),
            "Grand Total": [1.0] * len(vendors),
            "month filter": ["(multiple items)"] * len(vendors),
        }
    )
    answer_df = student_df.copy()

    result = grade_q10(student_df, answer_df, question_cfg={})

    assert result["structural_score"] == 1.0
    assert result["value_score"] == 1.0


def test_q10_grade_question_reports_vendor_count_failure_message() -> None:
    student_df = pd.DataFrame(
        {
            "vendor_name": _Q10_VENDORS[:4],
            "Honey": [0.1, 0.2, 0.3, 0.4],
            "No Promo Code": [0.9, 0.8, 0.7, 0.6],
            "month filter": ["(multiple items)"] * 4,
        }
    )
    answer_df = pd.DataFrame(
        {
            "vendor_name": _Q10_VENDORS,
            "Honey": [0.05] * len(_Q10_VENDORS),
            "No Promo Code": [0.95] * len(_Q10_VENDORS),
            "month filter": ["(multiple items)"] * len(_Q10_VENDORS),
        }
    )

    result = grade_q10(student_df, answer_df, question_cfg={})

    assert result["structural_score"] == 0.0
    assert any("Expected 14 vendors" in s for s in result["structural_issues"])


def test_evaluate_highlight_formatting_uses_workbook_fallback(tmp_path: Path) -> None:
    path, _ = _make_highlight_workbook(tmp_path, highlighted_rows=[3])
    score, issues = evaluate_highlight_formatting(path, None)
    assert score == 1.0
    assert issues == ["NEEDS_REVIEW: highlight check skipped"]


def test_evaluate_highlight_formatting_missing_highlight_returns_half_credit(tmp_path: Path) -> None:
    path, sname = _make_highlight_workbook(tmp_path, highlighted_rows=[])
    score, issues = evaluate_highlight_formatting(path, sname)
    assert score == 0.5
    assert issues == ["Missing highlight"]
